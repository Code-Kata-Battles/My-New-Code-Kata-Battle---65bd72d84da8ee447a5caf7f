import pandas as pd
from os import path

#Devuelve el nombre de la variable (sin los brackets)
def nombre_var(gvar):
    nombre = gvar.varName
    for cont, i in enumerate(nombre):
        if i == '[':
            return nombre[:cont]
        
#Devuelve los parametros de la variable en una lista
def parametro_var(gvar):
    nombre = gvar.varName
    contt = 0
    param = []
    for cont, i in enumerate(nombre):
        if i == '[':
            contt = cont
        if i in '],':
            param.append(nombre[contt+1:cont])
            contt = cont
    return param

#Ordena las variables en diccionarios
def ordenar(lista, cont):
    dic = {}
    nombre = ''
    
    for i in lista:
        param = parametro_var(i)
        if param[cont] == nombre:
            dic[param[cont]].append(i)
        else:
            nombre = param[cont]
            dic[nombre] = []
            dic[nombre].append(i)
            
    for i in dic:
        if len(dic[i]) > 1:
            dic[i] = ordenar(dic[i], cont + 1)
        else:
            dic[i] = dic[i][0].X
    return dic

#ejecucion principal
def ordenar_variables(var):
    
    #Guarda parametros necesarios
    pm = pd.read_excel("parametros.xlsx")
    
    categorias = pm.iloc[1, 2:15].values.tolist()
    categorias[9] = 'traje bano'
    bodegas = pm.iloc[134, 3:8].values.tolist()
    tiendas = pm.iloc[31:43, 3].values.tolist()
    periodos = [1, 2, 3, 4, 5, 6]
    puntos = bodegas + tiendas
    
    var_ordenadas = {}
    
    #Separa las variables en diccionarios, todo junto en una variable, var_ordenadas
    nombre = ''
    for i in var:
        if nombre_var(i) == nombre:
            var_ordenadas[nombre].append(i)
        else:
            nombre = nombre_var(i)
            var_ordenadas[nombre] = []
            var_ordenadas[nombre].append(i)
    for i in var_ordenadas:
        var_ordenadas[i] = ordenar(var_ordenadas[i], 0)
    
    #Pasa las variables desde los diccionarios a listas y las guarda en Excels ubicados en variables/
    
    #Variale X
    dfx = {}
    for t in periodos:
        dfx[t] = pd.DataFrame([[var_ordenadas['X'][categoria][punto][str(t)] for categoria in categorias] for punto in puntos],
                              index=[punto for punto in puntos],
                              columns=[categoria for categoria in categorias])
    with pd.ExcelWriter(path.join('variables', 'X.xlsx')) as writer:
        for t in periodos:
            dfx[t].to_excel(writer, sheet_name=f'Periodo {t}')
    
    #Variale Q
    dfq = {}
    for t in periodos:
        dfq[t] = pd.DataFrame([[var_ordenadas['Q'][categoria][punto][str(t)] for categoria in categorias] for punto in puntos],
                              index=[punto for punto in puntos],
                              columns=[categoria for categoria in categorias])
    with pd.ExcelWriter(path.join('variables', 'Q.xlsx') ) as writer:
        for t in periodos:
            dfq[t].to_excel(writer, sheet_name=f'Periodo {t}')
    
    #Variale GG
    dfgg = {}
    for t in periodos:
        dfgg[t] = pd.DataFrame([[var_ordenadas['GG'][categoria][bodega][str(t)] for categoria in categorias] for bodega in bodegas],
                              index=[bodega for bodega in bodegas],
                              columns=[categoria for categoria in categorias])
    with pd.ExcelWriter(path.join('variables', 'GG.xlsx') ) as writer:
        for t in periodos:
            dfgg[t].to_excel(writer, sheet_name=f'Periodo {t}')
    
    #Variale QC
    dfqc = {}
    for t in periodos:
        dfqc[t] = pd.DataFrame([[var_ordenadas['QC'][punto][bodega][str(t)] for punto in puntos] for bodega in bodegas],
                              index=[bodega for bodega in bodegas],
                              columns=[punto for punto in puntos])
    with pd.ExcelWriter(path.join('variables', 'QC.xlsx')) as writer:
        for t in periodos:
            dfqc[t].to_excel(writer, sheet_name=f'Periodo {t}')
    
    #Variale QCR
    dfqcr = {}
    dfqcr = pd.DataFrame([[var_ordenadas['QCR'][punto][str(periodo)] for punto in puntos] for periodo in periodos],
                          index=['Periodo' + str(periodo) for periodo in periodos],
                          columns=[punto for punto in puntos])
    with pd.ExcelWriter(path.join('variables', 'QCR.xlsx')) as writer:
        dfqcr.to_excel(writer)
    
    #Variale QCQ
    dfqcq = {}
    dfqcq = pd.DataFrame([[var_ordenadas['QCQ'][punto][str(periodo)] for punto in puntos] for periodo in periodos],
                          index=['Periodo' + str(periodo) for periodo in periodos],
                          columns=[punto for punto in puntos])
    with pd.ExcelWriter(path.join('variables', 'QCQ.xlsx')) as writer:
        dfqcq.to_excel(writer)
        
    #Variale G
    dfg = {}
    existe = True
    for k, bodega in enumerate(bodegas):
        for t in periodos:    
            dfg[t] = [[var_ordenadas['G'][categoria][punto][bodega][str(t)] for categoria in categorias] for punto in puntos]
                            
            dfg[t].append(['' for i in categorias])
            dfg[t].insert(0, ['Hacia:', bodega] + ['' for i in categorias[:-2]])
            
            dfg[t] = pd.DataFrame(dfg[t], 
                                  index=[punto for punto in [''] + puntos + ['']],
                                  columns=[categoria for categoria in categorias])
        
        if existe:
            with pd.ExcelWriter(path.join('variables', 'G.xlsx') ) as writer:
                for t in periodos:
                    dfg[t].to_excel(writer, sheet_name=f'Periodo {t}')
            existe = False
        else:
            for t in periodos:
                append_df_to_excel(path.join('variables', 'G.xlsx'), dfg[t], sheet_name=f'Periodo {t}', startrow=20*(k))
                
            

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
        